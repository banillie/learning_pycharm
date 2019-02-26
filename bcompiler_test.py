from openpyxl import Workbook
import datetime
from bcompiler.utils import project_data_from_master
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font

'''Function filters all reported milestones into dictionaries for passing into the time_difference function'''

def milestone_dict(name, dictionary, m_list, m_list2, title):
    """

    :type m_list: object
    """
    m_dict = {}
    if name in dictionary.keys():
        m_dict[title] = {}
        z = dictionary[name]
        # m_list is a list of milestone keys. The below loop puts
        # specific project milestone keys into a list.
        milestone_keys = []
        for x in m_list:
            a = z[x]
            milestone_keys.append(a)

        # m_list2 is a list of milestone dates. The below loop puts
        # specific project milestone dates into a list.
        milestone_dates = []
        for x in m_list2:
            b = z[x]
            milestone_dates.append(b)

        # the below loop places the above lists into dictionary.
        for i in range(0, len(milestone_keys)):
            milestone_key = milestone_keys[i]
            milestone_date = milestone_dates[i]
            # checks for dates being reported as strings. Does not handle
            # error but prints the problemtic data so can be changed in master
            # TODO - ideally the programme would handle this issue
            if type(milestone_date) == str:
                print(milestone_key, milestone_date)
            # handles none type reporting for milestone dates. These are not
            # needed in dictionaries.
            if milestone_date != None:
                m_dict[title][milestone_key] = milestone_date
            else:
                pass

    else:
        m_dict[name] = {}

    return m_dict


'''Function for calculating the time difference between milestone date and
the first day of interest (day 0 on graph that is produced). The date is set at a global level below.
This function also filters milestones between the date range of interest - also set below'''


def time_difference(title, dictionary):
    td_dict = {}
    td_dict[title] = {}
    for milestone in dictionary[title]:
        milestone_date = dictionary[title][milestone]
        if date_1 <= milestone_date <= date_2:  # milestone filtered with in selected range
            time_delta = (milestone_date - date_1).days  # time_delta calculated here
            td_dict[title][milestone] = time_delta

    return td_dict


'''function places data contained in dictionaries into excle spreadsheet
the 'place' information relates to where different types of milestones should be placed
in the spreadsheet. This is caluclated in the running fuction below'''


def placing_series_info(place, dictionary, dict_td, title, ws, data_label):
    row_in = 0
    for milestone in dict_td[title]:
        ws.cell(row=row_in + place, column=1, value=title)
        ws.cell(row=row_in + place, column=2, value=milestone)
        ws.cell(row=row_in + place, column=3, value=dictionary[title][milestone])
        ws.cell(row=row_in + place, column=4, value=dict_td[title][milestone])

        if title == 'approval':
            ws.cell(row=row_in + place, column=5, value=int('1'))
        elif title == 'project':
            ws.cell(row=row_in + place, column=5, value=int('2'))
        elif title == 'assurance':
            ws.cell(row=row_in + place, column=5, value=int('3'))

        for milestone_label in data_label:
            if milestone_label[0] == milestone:
                ws.cell(row=row_in + place, column=6, value=data_label.index(milestone_label) + 1)

        row_in += 1

    ws.cell(row=1, column=1, value='Type')
    ws.cell(row=1, column=2, value='Milestone')
    ws.cell(row=1, column=3, value='Date')
    ws.cell(row=1, column=4, value='Time Delta')
    ws.cell(row=1, column=5, value='Index')
    ws.cell(row=1, column=6, value='data_label')

    return ws


'''function that produces the chart'''


def build_chart(ws, name, project_place, assurance_place, total_place):
    approval_point = current_Q_dict[name]['BICC approval point']

    chart = ScatterChart()
    chart.title = str(name) + ' Schedule\n Last BICC Approved Business Case: ' + str(approval_point)
    chart.style = 18
    chart.x_axis.title = x_axis_title  # x_axis title set at global level below
    chart.y_axis.title = 'Milestones'
    chart.auto_axis = False
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = chart_x_axis_max  # max number (of days) in the x axis. set at global level below
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 4  # currently hard coded. TODO- amend so can be adapted to number of different types of milestone.
    chart.height = 9  # default is 7.5
    chart.width = 21  # default is 15

    '''changes units on x and y axis'''
    chart.x_axis.majorUnit = chart_x_major_unit  # this is the intervals for days - set at global level below.
    chart.y_axis.majorUnit = 1.0

    '''reverses y axis'''
    chart.x_axis.scaling.orientation = "minMax"
    chart.y_axis.scaling.orientation = "maxMin"

    '''makes the x axis cross at the max y value'''
    chart.x_axis.crosses = 'max'

    '''removes lable on y axis'''
    chart.y_axis.delete = True

    '''styling chart'''
    '''formating for titles'''
    font = Font(typeface='Calibri')
    size = 1200  # 12 point size
    cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.x_axis.title.tx.rich.p[0].pPr = pp  # x_axis title

    size_2 = 1400
    cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    pp_2 = ParagraphProperties(defRPr=cp_2)
    rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    chart.title.tx.rich.p[0].pPr = pp_2  # chart title

    '''the below assigns series information to the data that has been placed in the chart. 
    old values are placed first show that they show behind the current values'''

    xvalues = Reference(ws, min_col=4, min_row=2, max_row=project_place - 1)
    yvalues = Reference(ws, min_col=5, min_row=2, max_row=project_place - 1)
    series = Series(values=yvalues, xvalues=xvalues, title="Approval milestones")
    chart.series.append(series)
    s1 = chart.series[0]
    s1.marker.symbol = "diamond"
    s1.marker.size = 10
    s1.marker.graphicalProperties.solidFill = "c9e243"  # Marker filling greenish
    s1.marker.graphicalProperties.line.solidFill = "c9e243"  # Marker outline greenish
    s1.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=4, min_row=project_place, max_row=assurance_place - 1)
    yvalues = Reference(ws, min_col=5, min_row=project_place, max_row=assurance_place - 1)
    series = Series(values=yvalues, xvalues=xvalues, title="Project milestones")
    chart.series.append(series)
    s1 = chart.series[1]
    s1.marker.symbol = "circle"
    s1.marker.size = 10
    s1.marker.graphicalProperties.solidFill = "ff8c00"  # Marker filling orange
    s1.marker.graphicalProperties.line.solidFill = "ff8c00"  # Marker outline
    s1.graphicalProperties.line.noFill = True

    xvalues = Reference(ws, min_col=4, min_row=assurance_place, max_row=total_place - 1)
    yvalues = Reference(ws, min_col=5, min_row=assurance_place, max_row=total_place - 1)
    series = Series(values=yvalues, xvalues=xvalues, title="Assurance milestones")
    chart.series.append(series)
    s12 = chart.series[2]
    s12.marker.symbol = "square"
    s12.marker.size = 10
    s12.marker.graphicalProperties.solidFill = "68A900"  # Marker filling ??
    s12.marker.graphicalProperties.line.solidFill = "68A900"  # Marker outline ??
    s12.graphicalProperties.line.noFill = True

    ws.add_chart(chart, "J3")

    return ws


'''function calculates the data lable number for projects on a individual basis i.e.
the numbers revert back to one each time for each project'''


# TODO find a better fix than planning blank lists in for those type of milestones not being put in graph
def data_labels_singluar(project_name, approval_dict, project_dict, assurance_dict):
    label_dict = {}

    approval = approval_dict['approval']
    approval_list = list(approval.items())
    project = project_dict['project']
    project_list = list(project.items())
    assurance = assurance_dict['assurance']
    assurance_list = list(assurance.items())

    # assurance_list = [] # blank because not including assurance milestones in graph
    # project_list = [] #blanked to not include project milestones in graph

    all_in = approval_list + project_list + assurance_list

    all_in.sort(key=lambda x: x[1])

    return all_in


'''function that runs the programme'''


def running(proj_name, dict_1):
    print(proj_name)
    wb = Workbook()
    ws = wb.active
    '''firstly compiles dictionaries for each milestone group '''
    approval = milestone_dict(proj_name, dict_1, approval_milestone_keys, approval_milestone_dates, 'approval')
    project = milestone_dict(proj_name, dict_1, project_milestone_keys, project_milestone_dates, 'project')
    assurance = milestone_dict(proj_name, dict_1, assurance_milestone_keys, assurance_milestone_dates, 'assurance')

    '''secondly milestones are filtered and time_deltas are calculated and
    put into time delta (td) dictionaries'''
    approval_td = time_difference('approval', approval)
    project_td = time_difference('project', project)
    assurance_td = time_difference('assurance', assurance)

    '''data labels are calculated'''
    d_labels = data_labels_singluar(proj_name, approval_td, project_td, assurance_td)
    # print(d_labels)

    '''calculates where to place information into the spreadsheet'''
    approval_place = 2
    if len(approval_td['approval']) > 0:
        project_place = 2 + len(approval_td['approval'])
    else:
        project_place = 3

    if len(project_td['project']) > 0:
        assurance_place = project_place + len(project_td['project'])
    else:
        assurance_place = project_place + 1
    if len(assurance_td['assurance']) > 0:
        total_place = assurance_place + len(assurance_td['assurance'])
    else:
        total_place = assurance_place + 1

    placing_series_info(approval_place, approval, approval_td, 'approval', ws, d_labels)
    placing_series_info(project_place, project, project_td, 'project', ws, d_labels)
    placing_series_info(assurance_place, assurance, assurance_td, 'assurance', ws, d_labels)

    build_chart(ws, x, project_place, assurance_place, total_place)

    return wb


'''lists containing milestone meta data of interest. These list include everything. However, they can be amended
if only certain milestones (e.g. just BICC business case approvals) are being interrogated'''

approval_milestone_keys = ['Approval MM1', 'Approval MM2', 'Approval MM3', 'Approval MM4', 'Approval MM5',
                           'Approval MM6',
                           'Approval MM7', 'Approval MM8', 'Approval MM9', 'Approval MM10', 'Approval MM11',
                           'Approval MM12',
                           'Approval MM13', 'Approval MM14', 'Approval MM15', 'Approval MM16']

project_milestone_keys = ['Project MM18', 'Project MM19', 'Project MM20', 'Project MM21', 'Project MM22',
                          'Project MM23',
                          'Project MM24', 'Project MM25', 'Project MM26', 'Project MM27', 'Project MM28',
                          'Project MM29',
                          'Project MM30', 'Project MM31', 'Project MM32']

assurance_milestone_keys = ['Assurance MM1', 'Assurance MM2', 'Assurance MM3', 'Assurance MM4', 'Assurance MM5',
                            'Assurance MM6',
                            'Assurance MM7', 'Assurance MM8', 'Assurance MM9', 'Assurance MM10', 'Assurance MM11',
                            'Assurance MM12',
                            'Assurance MM13', 'Assurance MM14', 'Assurance MM15', 'Assurance MM16', 'Assurance MM17',
                            'Assurance MM18']

approval_milestone_dates = ['Approval MM1 Forecast / Actual', 'Approval MM2 Forecast / Actual',
                            'Approval MM3 Forecast / Actual',
                            'Approval MM4 Forecast / Actual', 'Approval MM5 Forecast / Actual',
                            'Approval MM6 Forecast / Actual',
                            'Approval MM7 Forecast / Actual', 'Approval MM8 Forecast / Actual',
                            'Approval MM9 Forecast / Actual',
                            'Approval MM10 Forecast / Actual', 'Approval MM11 Forecast / Actual',
                            'Approval MM12 Forecast / Actual',
                            'Approval MM13 Forecast - Actual', 'Approval MM14 Forecast - Actual',
                            'Approval MM15 Forecast - Actual',
                            'Approval MM16 Forecast - Actual']

project_milestone_dates = ['Project MM18 Forecast - Actual', 'Project MM19 Forecast - Actual',
                           'Project MM20 Forecast - Actual',
                           'Project MM21 Forecast - Actual', 'Project MM22 Forecast - Actual',
                           'Project MM23 Forecast - Actual',
                           'Project MM24 Forecast - Actual', 'Project MM25 Forecast - Actual',
                           'Project MM26 Forecast - Actual',
                           'Project MM27 Forecast - Actual', 'Project MM28 Forecast - Actual',
                           'Project MM29 Forecast - Actual',
                           'Project MM30 Forecast - Actual', 'Project MM31 Forecast - Actual',
                           'Project MM32 Forecast - Actual']

assurance_milestone_dates = ['Assurance MM1 Forecast - Actual', 'Assurance MM2 Forecast - Actual',
                             'Assurance MM3 Forecast - Actual',
                             'Assurance MM4 Forecast - Actual', 'Assurance MM5 Forecast - Actual',
                             'Assurance MM6 Forecast - Actual',
                             'Assurance MM7 Forecast - Actual', 'Assurance MM8 Forecast - Actual',
                             'Assurance MM9 Forecast - Actual',
                             'Assurance MM10 Forecast - Actual', 'Assurance MM11 Forecast - Actual',
                             'Assurance MM12 Forecast - Actual',
                             'Assurance MM13 Forecast - Actual', 'Assurance MM14 Forecast - Actual',
                             'Assurance MM15 Forecast - Actual',
                             'Assurance MM16 Forecast - Actual', 'Assurance MM17 Forecast - Actual',
                             'Assurance MM18 Forecast - Actual']

'''get master dictionary for quarter of interest'''
current_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')

'''options for passing different combination of projects into programme'''
'''all projects in portfolio'''
current_Q_list = list(current_Q_dict.keys())

'''groups of projects'''
# current_Q_list = ['East Midlands Franchise','South Eastern Rail Franchise Competition', 'West Coast Partnership Franchise', 'Cross Country Rail Franchise Competition']      # rail franchising
'''one project'''
# current_Q_list = ['Heathrow Expansion']

'''Options to set scaling and format of chart'''
'''date range setting'''
date_1 = datetime.date(2018, 9, 1)  # sets the first date of interest
date_2 = datetime.date(2019, 10, 1)  # sets the second date of interest

'''chart styling information'''
x_axis_title = 'Sept 18 to Sept 19'  # insert x axis title here
chart_x_axis_max = (date_2 - date_1).days  # calculated automatically based on date range set above
chart_x_major_unit = 30

'''final code that calls the programme to run name of file
should be change to reflect the graph that is being returned'''
for x in current_Q_list:
    print_miles = running(x, current_Q_dict)
    print_miles.save('C:\\Users\\Standalone\\Will\\PC_test_Q2_1819_{}_milestone.xlsx'.format(x))